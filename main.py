"""
주소 입력 → 건축물 면적 조회 → 엑셀 출력 에이전트.

사용법:
    python3 main.py                              # input.xlsx 읽어 output.xlsx 생성
    python3 main.py --input my.xlsx --col 0      # 컬럼 인덱스 지정
"""

import argparse
import os
import re
import requests
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime
from dataclasses import dataclass
from collections import defaultdict
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# API 키는 환경변수 또는 Streamlit secrets에서 읽음. 코드에 하드코딩 X.
# 주소→PNU 변환은 VWorld 지오코딩 사용 (juso.go.kr는 해외 IP에서 timeout 발생).
DATA_GO_KR_KEY = os.environ.get("DATA_GO_KR_KEY", "")
VWORLD_API_KEY = os.environ.get("VWORLD_API_KEY", "")

VWORLD_GEOCODE_URL = "https://api.vworld.kr/req/address"
BR_BASE = "https://apis.data.go.kr/1613000/BldRgstHubService"


@dataclass
class AddressCode:
    sigungu_cd: str  # 5자리 시군구코드
    bjdong_cd: str   # 5자리 법정동코드
    bun: str         # 4자리 본번 (zero-padded)
    ji: str          # 4자리 부번 (zero-padded)
    plat_gb_cd: str  # 0=대지, 1=산
    road_addr: str   # 도로명 주소 (참고용)
    jibun_addr: str  # 정규화된 지번 주소 (참고용)


def _normalize_address_query(q: str) -> str:
    """juso.go.kr이 인식하기 좋은 형태로 주소 정규화.

    - "대구 광역시" → "대구광역시" (광역단위 명칭 사이 공백 제거)
    - 연속 공백 1개로
    """
    q = re.sub(r"\s+", " ", q).strip()
    for word in ("특별시", "광역시", "특별자치시", "특별자치도"):
        q = re.sub(rf"(\S)\s+{word}", rf"\1{word}", q)
    return q


def lookup_address(jibun_query: str) -> AddressCode | None:
    """지번 주소 → 건축물대장 API용 코드. VWorld 지오코딩 사용.

    juso.go.kr은 해외 IP에서 차단/timeout이 빈번해 클라우드 배포 환경에서 부적합.
    VWorld의 PNU(level4LC, 19자리)에서 sigunguCd/bjdongCd/본번/부번을 분해 추출.
    """
    if not VWORLD_API_KEY:
        raise RuntimeError("VWORLD_API_KEY 미설정 (.streamlit/secrets.toml 확인)")
    candidates: list[str] = []
    normalized = _normalize_address_query(jibun_query)
    candidates.append(normalized)
    if jibun_query.strip() != normalized:
        candidates.append(jibun_query.strip())

    for q in candidates:
        params = {
            "service": "address",
            "request": "getcoord",
            "type": "parcel",
            "address": q,
            "crs": "EPSG:4326",
            "format": "json",
            "key": VWORLD_API_KEY,
        }
        try:
            r = requests.get(VWORLD_GEOCODE_URL, params=params, timeout=15)
            r.raise_for_status()
            data = r.json().get("response", {})
        except (requests.RequestException, ValueError) as e:
            raise RuntimeError(f"VWorld 지오코딩 호출 실패: {e}")

        if data.get("status") != "OK":
            continue

        struct = data.get("refined", {}).get("structure", {})
        pnu = struct.get("level4LC", "") or ""
        if len(pnu) != 19 or not pnu.isdigit():
            continue
        # PNU 11번째 자리: '1'=일반(대지), '2'=산. platGbCd: 0=대지, 1=산.
        plat_gb = "1" if pnu[10] == "2" else "0"
        return AddressCode(
            sigungu_cd=pnu[:5],
            bjdong_cd=pnu[5:10],
            bun=pnu[11:15],
            ji=pnu[15:19],
            plat_gb_cd=plat_gb,
            road_addr="",  # VWorld 지오코딩에 도로명주소는 별도. 미리보기에서 빈값 표시.
            jibun_addr=data.get("refined", {}).get("text", q),
        )
    return None


@dataclass
class BuildingTitle:
    """표제부 1동(棟) 분 정보."""
    bld_nm: str
    dong_nm: str
    regstr_gb: str           # "일반" or "집합"
    plat_area: float         # 대지면적 (㎡)
    arch_area: float         # 건축면적
    tot_area: float          # 연면적
    vl_rat_estm_tot_area: float  # 용적률산정용 연면적
    main_purps: str          # 주용도
    grnd_flr_cnt: int        # 지상층수
    ugrnd_flr_cnt: int       # 지하층수
    use_apr_day: datetime | None  # 사용승인일


def _to_float(v) -> float:
    try: return float(v)
    except (TypeError, ValueError): return 0.0


def _to_int(v) -> int:
    try: return int(float(v))
    except (TypeError, ValueError): return 0


def _parse_yyyymmdd(v) -> datetime | None:
    """건축물대장 API의 8자리 일자 문자열을 datetime으로 변환. 빈값/이상값은 None."""
    if v is None: return None
    s = str(v).strip()
    if len(s) != 8 or not s.isdigit():
        return None
    try:
        return datetime.strptime(s, "%Y%m%d")
    except ValueError:
        return None


def get_title_info(addr: AddressCode) -> list[BuildingTitle]:
    """건축물대장 표제부 호출. 같은 지번에 속한 모든 동을 반환."""
    if not DATA_GO_KR_KEY:
        raise RuntimeError("DATA_GO_KR_KEY 미설정 (.streamlit/secrets.toml 확인)")
    params = {
        "serviceKey": DATA_GO_KR_KEY,
        "sigunguCd": addr.sigungu_cd,
        "bjdongCd": addr.bjdong_cd,
        "bun": addr.bun,
        "ji": addr.ji,
        "platGbCd": addr.plat_gb_cd,
        "numOfRows": 100,
        "pageNo": 1,
        "_type": "json",
    }
    r = requests.get(f"{BR_BASE}/getBrTitleInfo", params=params, timeout=15)
    r.raise_for_status()
    body = r.json()["response"]["body"]
    raw = body.get("items", {}).get("item", [])
    if not raw: return []
    raw = raw if isinstance(raw, list) else [raw]

    titles = [
        BuildingTitle(
            bld_nm=it.get("bldNm", "") or "",
            dong_nm=it.get("dongNm", "") or "",
            regstr_gb=it.get("regstrGbCdNm", "") or "",
            plat_area=_to_float(it.get("platArea")),
            arch_area=_to_float(it.get("archArea")),
            tot_area=_to_float(it.get("totArea")),
            vl_rat_estm_tot_area=_to_float(it.get("vlRatEstmTotArea")),
            main_purps=it.get("mainPurpsCdNm", "") or "",
            grnd_flr_cnt=_to_int(it.get("grndFlrCnt")),
            ugrnd_flr_cnt=_to_int(it.get("ugrndFlrCnt")),
            use_apr_day=_parse_yyyymmdd(it.get("useAprDay")),
        )
        for it in raw
    ]

    # 단지(다중 동)일 때 대지면적이 1개 동에만 입력되는 패턴 보정.
    # 같은 지번의 max(platArea)로 0인 동들을 채워 모두 단지 전체 대지면적을 갖도록 함.
    max_plat = max((t.plat_area for t in titles), default=0.0)
    if max_plat > 0:
        for t in titles:
            if t.plat_area == 0:
                t.plat_area = max_plat
    return titles


@dataclass
class BuildingUnit:
    """집합건물의 한 호(號) 분 정보."""
    dong_nm: str
    ho_nm: str
    main_purps: str          # 호 주용도 (아파트/오피스텔/근린생활시설 등)
    exclusive_area: float    # 전용면적 합 (㎡)
    common_area: float       # 공용면적 합 (㎡)
    floor_label: str         # 층 정보 (예: "61층")


def _fetch_unit_page(addr: AddressCode, page: int) -> dict:
    """전유공용면적 API 한 페이지 호출. 항상 numOfRows=100 (API가 무시하고 100으로 캡)."""
    params = {
        "serviceKey": DATA_GO_KR_KEY,
        "sigunguCd": addr.sigungu_cd,
        "bjdongCd": addr.bjdong_cd,
        "bun": addr.bun,
        "ji": addr.ji,
        "platGbCd": addr.plat_gb_cd,
        "numOfRows": 100,
        "pageNo": page,
        "_type": "json",
    }
    r = requests.get(f"{BR_BASE}/getBrExposPubuseAreaInfo", params=params, timeout=30)
    r.raise_for_status()
    body = r.json()["response"]["body"]
    items = body.get("items", {}).get("item", []) or []
    items = items if isinstance(items, list) else [items]
    return {"items": items, "totalCount": int(body.get("totalCount", 0) or 0)}


def get_unit_areas(addr: AddressCode, verbose: bool = False,
                   on_progress=None) -> list[BuildingUnit]:
    """getBrExposPubuseAreaInfo 호출. 지상에 위치한 호만 반환, 전유/공용 합산.

    응답 1호 = 여러 행 (전유 1행 + 공용 N행)이라 (동, 호)로 그룹화 후 area 합산.
    1페이지 호출로 totalCount 파악 후 나머지 페이지를 ThreadPool로 병렬 호출.

    on_progress(done, total): 누적 행 수와 totalCount를 받는 콜백 (Streamlit 진행률용).
    """
    first = _fetch_unit_page(addr, 1)
    rows: list[dict] = list(first["items"])
    total = first["totalCount"]
    if verbose:
        print(f"    페이지 1/{(total + 99) // 100} 누적 {len(rows):,}/{total:,}",
              end="\r", flush=True)
    if on_progress:
        on_progress(len(rows), max(total, 1))

    if total > 100:
        n_pages = (total + 99) // 100
        with ThreadPoolExecutor(max_workers=10) as ex:
            for body in ex.map(lambda p: _fetch_unit_page(addr, p),
                                range(2, n_pages + 1)):
                rows.extend(body["items"])
                if verbose:
                    print(f"    누적 {len(rows):,}/{total:,}", end="\r", flush=True)
                if on_progress:
                    on_progress(len(rows), max(total, 1))
    if verbose:
        print()

    groups: dict[tuple[str, str], list[dict]] = defaultdict(list)
    for it in rows:
        groups[(it.get("dongNm", "") or "", it.get("hoNm", "") or "")].append(it)

    units: list[BuildingUnit] = []
    for (dong, ho), entries in groups.items():
        exclusive = [e for e in entries if e.get("exposPubuseGbCdNm") == "전유"]
        commons = [e for e in entries if e.get("exposPubuseGbCdNm") == "공용"]
        if not exclusive:
            continue
        # 호의 위치층 = 전유 행 기준. 지상층 호만 포함 (사용자 요구: 지상 건축물).
        if exclusive[0].get("flrGbCdNm") != "지상":
            continue
        # 위치층 라벨: flrNo + flrGbCdNm 조합 우선 (일부 지자체에서 flrNoNm에
        # 호명이 잘못 들어와 있는 사례가 있어 신뢰 불가).
        flr_no = _to_int(exclusive[0].get("flrNo"))
        flr_gb = exclusive[0].get("flrGbCdNm", "") or ""
        if flr_no > 0 and flr_gb in ("지상", "지하"):
            floor_label = f"{flr_gb} {flr_no}층"
        else:
            floor_label = exclusive[0].get("flrNoNm", "") or ""
        units.append(BuildingUnit(
            dong_nm=dong,
            ho_nm=ho,
            main_purps=exclusive[0].get("mainPurpsCdNm", "") or "",
            exclusive_area=sum(_to_float(e.get("area")) for e in exclusive),
            common_area=sum(_to_float(e.get("area")) for e in commons),
            floor_label=floor_label,
        ))

    units.sort(key=lambda u: (u.dong_nm, u.ho_nm))
    return units


# 사용자 양식의 4개 카테고리를 그룹 헤더(병합)로, 그 아래 세부 컬럼을 둠.
# (col_idx, group, label) 순. 1-indexed. group=None이면 일반 단일 헤더.
COL_SPEC = [
    ("번호",       None,            "번호"),
    ("입력주소",   None,            "입력주소"),
    ("도로명주소", None,            "도로명주소"),
    ("건물명",     "건축인허가기본개요", "건물명"),
    ("동명",       "건축인허가기본개요", "동명"),
    ("주용도",     "건축인허가기본개요", "주용도"),
    ("지상층수",   "건축인허가기본개요", "지상층수"),
    ("지하층수",   "건축인허가기본개요", "지하층수"),
    ("대지면적",   "건축인허가전유공용면적", "대지면적(㎡)"),
    ("건축면적",   "건축인허가전유공용면적", "건축면적(㎡)"),
    ("연면적",     "건축인허가전유공용면적", "연면적(㎡)"),
    ("용적률산정", "건축인허가전유공용면적", "용적률산정용연면적(㎡)"),
    ("호명",       "호별전유공용면적", "호명"),
    ("위치층",     "호별전유공용면적", "위치층"),
    ("전용면적",   "호별전유공용면적", "전용면적(㎡)"),
    ("공용면적",   "호별전유공용면적", "공용면적(㎡)"),
    ("주택유형",   None,            "주택유형"),
    ("사용승인일", None,            "사용승인일"),
    ("비고",       None,            "비고"),
]
COL_WIDTHS = [6, 32, 38, 18, 14, 16, 10, 10, 12, 12, 12, 18, 14, 12, 12, 12, 18, 14, 28]


def _row(no, addr_in, road, t: "BuildingTitle | None", u: "BuildingUnit | None", note: str) -> list:
    """COL_SPEC 순서에 맞춘 19개 셀 한 줄 생성."""
    bld = t.bld_nm if t else ""
    dong = t.dong_nm if t else ""
    main_purps_t = t.main_purps if t else ""
    grnd = t.grnd_flr_cnt if t else ""
    ugrnd = t.ugrnd_flr_cnt if t else ""
    plat = t.plat_area if t else ""
    arch = t.arch_area if t else ""
    tot = t.tot_area if t else ""
    vlrat = t.vl_rat_estm_tot_area if t else ""
    ho = u.ho_nm if u else ""
    floor = u.floor_label if u else ""
    excl = u.exclusive_area if u else ""
    comm = u.common_area if u else ""
    house_type = (u.main_purps if u else "") or main_purps_t
    use_apr = (t.use_apr_day if t else None) or ""
    return [no, addr_in, road, bld, dong, main_purps_t, grnd, ugrnd,
            plat, arch, tot, vlrat, ho, floor, excl, comm, house_type, use_apr, note]


def process_address(no, query: str, verbose: bool = True,
                    on_progress=None) -> list[list]:
    """한 주소 → 출력 행 리스트. 에러여도 최소 1행은 반환(비고에 사유).

    on_progress(done, total): 호별 면적 페이징 진행률 콜백. 집합건물에서만 호출됨.
    """
    try:
        addr = lookup_address(query)
    except Exception as e:
        return [_row(no, query, "", None, None, f"주소 변환 실패: {e}")]
    if addr is None:
        return [_row(no, query, "", None, None, "주소 검색 결과 없음")]

    try:
        titles = get_title_info(addr)
    except Exception as e:
        return [_row(no, query, addr.road_addr, None, None, f"표제부 호출 실패: {e}")]
    if not titles:
        return [_row(no, query, addr.road_addr, None, None, "건축물대장 정보 없음")]

    units_by_dong: dict[str, list[BuildingUnit]] = {}
    if any(t.regstr_gb == "집합" for t in titles):
        try:
            units = get_unit_areas(addr, verbose=verbose, on_progress=on_progress)
            for u in units:
                units_by_dong.setdefault(u.dong_nm, []).append(u)
        except Exception as e:
            return [_row(no, query, addr.road_addr, None, None, f"호별 정보 호출 실패: {e}")]

    rows: list[list] = []
    for t in titles:
        dong_units = units_by_dong.get(t.dong_nm, [])
        if t.regstr_gb == "집합" and dong_units:
            for u in dong_units:
                rows.append(_row(no, query, addr.road_addr, t, u, ""))
        else:
            note = "" if t.regstr_gb == "일반" else "호별 데이터 없음"
            rows.append(_row(no, query, addr.road_addr, t, None, note))
    return rows


def read_input(input_path: str) -> list[tuple]:
    """샘플 양식 (번호 | 주소 | ...) 첫 시트를 읽어 [(no, address), ...] 반환."""
    wb = load_workbook(input_path, read_only=True, data_only=True)
    ws = wb.active
    out: list[tuple] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if not row or len(row) < 2 or row[1] is None:
            continue
        addr = str(row[1]).strip()
        if not addr:
            continue
        no = row[0] if row[0] is not None else (i)
        out.append((no, addr))
    wb.close()
    return out


def write_output(rows: list[list], output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "조회결과"

    # --- 2단 병합 헤더 ---
    # 행1: 그룹 헤더 (또는 단일 헤더면 행1~2 병합)
    # 행2: 세부 헤더
    groups: list[tuple[str | None, int, int]] = []  # (group, start, end)
    i = 0
    while i < len(COL_SPEC):
        g = COL_SPEC[i][1]
        j = i
        while j + 1 < len(COL_SPEC) and COL_SPEC[j + 1][1] == g and g is not None:
            j += 1
        groups.append((g, i + 1, j + 1))
        i = j + 1

    header_fill = PatternFill("solid", fgColor="DDEBF7")
    bold_center = {"font": Font(bold=True),
                   "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
                   "fill": header_fill}

    for g, start, end in groups:
        if g is None:
            cell = ws.cell(row=1, column=start, value=COL_SPEC[start - 1][2])
            for k, v in bold_center.items(): setattr(cell, k, v)
            ws.merge_cells(start_row=1, start_column=start, end_row=2, end_column=start)
        else:
            cell = ws.cell(row=1, column=start, value=g)
            for k, v in bold_center.items(): setattr(cell, k, v)
            ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
            for c in range(start, end + 1):
                sub = ws.cell(row=2, column=c, value=COL_SPEC[c - 1][2])
                for k, v in bold_center.items(): setattr(sub, k, v)

    # --- 데이터 ---
    num_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")
    for r_offset, row in enumerate(rows, start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_offset, column=c_idx, value=value)
            if isinstance(value, datetime):
                cell.number_format = "yyyy-mm-dd"
                cell.alignment = center_align
            elif isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = num_align

    # 너비/고정/행높이
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 22
    ws.freeze_panes = "D3"
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="주소 → 건축물 면적 → 엑셀 출력")
    parser.add_argument("--input", default="조회대상샘플.xlsx")
    parser.add_argument("--output", default="조회결과.xlsx")
    args = parser.parse_args()

    entries = read_input(args.input)
    print(f"입력 주소 {len(entries)}건. 처리 시작...\n")

    all_rows: list[list] = []
    for i, (no, q) in enumerate(entries, 1):
        print(f"[{i}/{len(entries)}] No.{no}  {q}")
        rows = process_address(no, q)
        all_rows.extend(rows)
        print(f"  → {len(rows)}행 생성\n")

    write_output(all_rows, args.output)
    print(f"완료: {args.output} (데이터 {len(all_rows)}행)")


if __name__ == "__main__":
    main()
