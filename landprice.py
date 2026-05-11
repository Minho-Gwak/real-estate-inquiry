"""개별공시지가 조회 라이브러리.

주소 → PNU(19자리) → 최근 3개년 공시지가 (원/㎡).

사용 API:
  - juso.go.kr 도로명주소 검색 → admCd + 본/부번 + 산여부 → PNU 조립
  - VWorld NED 개별공시지가 (api.vworld.kr/ned/data/getIndvdLandPriceAttr)

환경변수:
  - JUSO_API_KEY      (main.py와 공유)
  - VWORLD_API_KEY    (개별공시지가 전용)

이 파일의 함수들은 라이브러리로 호출 가능. CLI는 별도 두지 않음.
"""

from __future__ import annotations

import datetime as dt
import os
import re
import time
from concurrent.futures import ThreadPoolExecutor
from dataclasses import dataclass
from xml.etree import ElementTree as ET

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

VWORLD_GEOCODE_URL = "https://api.vworld.kr/req/address"
# 토지특성정보: 가격(pblntfPclnd)과 면적(lndpclAr)을 한 번에 반환.
LAND_ATTR_ENDPOINT = "https://api.vworld.kr/ned/data/getLandCharacteristics"
TIMEOUT_SEC = 15
# VWorld 키에 등록된 도메인. 로컬은 localhost, 배포 시 env로 주입.
VWORLD_DOMAIN = os.environ.get("VWORLD_DOMAIN", "localhost")

# 한국 공공 API가 default Python UA를 차단하는 사례 회피용 브라우저 UA.
HTTP_HEADERS = {
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                  "AppleWebKit/537.36 (KHTML, like Gecko) "
                  "Chrome/124.0.0.0 Safari/537.36",
    "Accept": "application/json, text/plain, */*",
}


def _normalize_query(q: str) -> str:
    """광역시·특별시 띄어쓰기 정규화 (main.py와 동일 로직)."""
    q = re.sub(r"\s+", " ", q).strip()
    for word in ("특별시", "광역시", "특별자치시", "특별자치도"):
        q = re.sub(rf"(\S)\s+{word}", rf"\1{word}", q)
    return q


@dataclass
class PnuResolution:
    pnu: str | None
    matched_jibun: str | None
    status: str       # "ok" | "not_found" | "api_error"
    error_msg: str = ""


def resolve_pnu(query: str) -> PnuResolution:
    """주소 문자열 → PNU. VWorld 지오코딩 사용 (해외 IP에서도 안정).

    VWorld 응답의 refined.structure.level4LC가 PNU 19자리 그대로.
    """
    api_key = os.environ.get("VWORLD_API_KEY", "").strip()
    if not api_key:
        return PnuResolution(None, None, "api_error", "VWORLD_API_KEY 미설정")

    candidates: list[str] = []
    norm = _normalize_query(query)
    candidates.append(norm)
    if query.strip() != norm:
        candidates.append(query.strip())

    last_msg = ""
    for q in candidates:
        try:
            r = requests.get(
                VWORLD_GEOCODE_URL,
                params={
                    "service": "address",
                    "request": "getcoord",
                    "type": "parcel",
                    "address": q,
                    "crs": "EPSG:4326",
                    "format": "json",
                    "domain": VWORLD_DOMAIN,
                    "key": api_key,
                },
                headers=HTTP_HEADERS,
                timeout=TIMEOUT_SEC,
            )
            r.raise_for_status()
            data = r.json().get("response", {})
        except (requests.RequestException, ValueError) as e:
            return PnuResolution(None, None, "api_error", f"VWorld 지오코딩 호출 실패: {e}")

        if data.get("status") != "OK":
            last_msg = f"검색 결과 없음 (status={data.get('status')})"
            continue

        struct = data.get("refined", {}).get("structure", {})
        pnu = struct.get("level4LC", "") or ""
        if len(pnu) != 19 or not pnu.isdigit():
            last_msg = f"PNU 형식 오류: {pnu!r}"
            continue
        return PnuResolution(
            pnu=pnu,
            matched_jibun=data.get("refined", {}).get("text", q),
            status="ok",
        )
    return PnuResolution(None, None, "not_found", last_msg or "검색 결과 없음")


def determine_target_years(today: dt.date | None = None) -> list[int]:
    """최근 3개년 = [올해-2, 올해-1, 올해]. 5/31 발표 이전이면 올해 값은 비어있을 수 있음."""
    if today is None:
        today = dt.date.today()
    return [today.year - 2, today.year - 1, today.year]


@dataclass
class LandAttr:
    """토지특성 한 건 응답."""
    price: float | None       # 공시지가 (원/㎡)
    area: float | None        # 토지 면적 (㎡)
    land_use: str | None      # 지목 (예: "대", "전")
    zone: str | None          # 용도지역 (예: "일반상업지역")
    status: str               # "ok" | "no_data" | "api_error: ..."


def _parse_land_attr_json(data: dict) -> LandAttr:
    container = data.get("landCharacteristicss") or data
    code = (container.get("resultCode") or "").strip()
    if code:
        return LandAttr(None, None, None, None,
                        f"api_error: {code} {(container.get('resultMsg') or '').strip()}".strip())
    fields = container.get("field")
    if not fields:
        return LandAttr(None, None, None, None, "no_data")
    if isinstance(fields, dict):
        fields = [fields]
    f = fields[0]

    def _num(raw):
        if raw is None or str(raw).strip() == "":
            return None
        try:
            return float(raw)
        except ValueError:
            return None

    def _str(raw):
        s = str(raw).strip() if raw is not None else ""
        return s or None

    price = _num(f.get("pblntfPclnd"))
    area = _num(f.get("lndpclAr"))
    land_use = _str(f.get("lndcgrCodeNm"))
    zone = _str(f.get("prposArea1Nm"))
    if price is None and area is None and land_use is None:
        return LandAttr(None, None, None, None, "no_data")
    return LandAttr(price, area, land_use, zone, "ok")


def _parse_land_attr_xml(text: str) -> LandAttr:
    try:
        root = ET.fromstring(text)
    except ET.ParseError as e:
        return LandAttr(None, None, None, None, f"api_error: xml parse {e}")
    err = root.find(".//returnAuthMsg") or root.find(".//errMsg")
    if err is not None and err.text:
        return LandAttr(None, None, None, None, f"api_error: {err.text}")

    def _num_node(name):
        n = root.find(f".//{name}")
        if n is None or not (n.text or "").strip():
            return None
        try:
            return float(n.text)
        except ValueError:
            return None

    def _str_node(name):
        n = root.find(f".//{name}")
        return (n.text or "").strip() if (n is not None and n.text) else None

    price = _num_node("pblntfPclnd")
    area = _num_node("lndpclAr")
    land_use = _str_node("lndcgrCodeNm")
    zone = _str_node("prposArea1Nm")
    if price is None and area is None and land_use is None:
        return LandAttr(None, None, None, None, "no_data")
    return LandAttr(price, area, land_use, zone, "ok")


def fetch_land_attr(pnu: str, year: int) -> LandAttr:
    """PNU + 연도 → LandAttr (가격·면적·지목·용도지역). VWorld 토지특성 NED 호출."""
    key = os.environ.get("VWORLD_API_KEY", "").strip()
    if not key:
        return LandAttr(None, None, None, None, "api_error: VWORLD_API_KEY 미설정")
    try:
        r = requests.get(
            LAND_ATTR_ENDPOINT,
            params={
                "key": key,
                "pnu": pnu,
                "stdrYear": str(year),
                "format": "json",
                "domain": VWORLD_DOMAIN,
                "numOfRows": "5",
                "pageNo": "1",
            },
            headers=HTTP_HEADERS,
            timeout=TIMEOUT_SEC,
        )
        r.raise_for_status()
        ct = r.headers.get("Content-Type", "")
        text = r.text
        if "xml" in ct.lower() or text.lstrip().startswith("<"):
            return _parse_land_attr_xml(text)
        return _parse_land_attr_json(r.json())
    except requests.RequestException as e:
        return LandAttr(None, None, None, None, f"api_error: network {e}")
    except ValueError as e:
        return LandAttr(None, None, None, None, f"api_error: parse {e}")


@dataclass
class LandPriceResult:
    query: str
    matched_jibun: str | None
    pnu: str | None
    years: list[int]
    prices: dict[int, float | None]   # 연도 → 가격(원/㎡) or None
    area: float | None                 # 토지 면적 (㎡)
    land_use: str | None               # 지목 (예: 대, 전, 답)
    zone: str | None                   # 용도지역 (예: 일반상업지역)
    status: str                        # "ok" | "not_found" | "no_data" | "api_error"
    error_msg: str = ""


def lookup_landprice(query: str, years: list[int] | None = None) -> LandPriceResult:
    """주소 → 최근 3개년 공시지가 + 면적·지목·용도지역. 연도별 호출은 병렬."""
    if years is None:
        years = determine_target_years()
    res = resolve_pnu(query)
    if res.status != "ok" or not res.pnu:
        return LandPriceResult(
            query=query,
            matched_jibun=res.matched_jibun,
            pnu=res.pnu,
            years=years,
            prices={y: None for y in years},
            area=None,
            land_use=None,
            zone=None,
            status=res.status,
            error_msg=res.error_msg,
        )

    prices: dict[int, float | None] = {}
    area: float | None = None
    land_use: str | None = None
    zone: str | None = None
    error_messages: list[str] = []
    with ThreadPoolExecutor(max_workers=3) as ex:
        for year, attr in zip(
            years, ex.map(lambda y: fetch_land_attr(res.pnu, y), years)
        ):
            prices[year] = attr.price
            if area is None and attr.area is not None:
                area = attr.area
            if land_use is None and attr.land_use:
                land_use = attr.land_use
            if zone is None and attr.zone:
                zone = attr.zone
            if attr.status.startswith("api_error"):
                error_messages.append(f"{year}: {attr.status}")

    status = "ok"
    if all(p is None for p in prices.values()) and area is None:
        status = "no_data"
    return LandPriceResult(
        query=query,
        matched_jibun=res.matched_jibun,
        pnu=res.pnu,
        years=years,
        prices=prices,
        area=area,
        land_use=land_use,
        zone=zone,
        status=status,
        error_msg="; ".join(error_messages),
    )


def write_landprice_xlsx(results: list[LandPriceResult], output_path) -> None:
    """개별공시지가 결과를 엑셀로 저장."""
    if not results:
        years = determine_target_years()
    else:
        years = results[0].years

    wb = Workbook()
    ws = wb.active
    ws.title = "개별공시지가"

    headers = ["번호", "입력주소", "매칭지번", "PNU", "지목", "용도지역", "면적(㎡)"] + \
              [f"{y}년 공시지가 (원/㎡)" for y in years] + ["비고"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="FFE5D9")  # OK 오렌지 톤
    bold_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = bold_center

    right_align = Alignment(horizontal="right", vertical="center")
    AREA_COL = 7                # 면적 컬럼 (지목·용도지역 추가로 +2)
    PRICE_START_COL = 8         # 첫 번째 가격 컬럼

    for i, r in enumerate(results, start=1):
        row: list = [i, r.query, r.matched_jibun or "", r.pnu or "",
                     r.land_use or "", r.zone or "", r.area]
        row.extend(r.prices[y] for y in years)
        note = ""
        if r.status == "not_found":
            note = r.error_msg or "주소 검색 실패"
        elif r.status == "api_error":
            note = r.error_msg or "API 오류"
        elif r.status == "no_data":
            note = "공시지가 데이터 없음"
        elif r.error_msg:
            note = r.error_msg
        row.append(note)
        ws.append(row)

        # 면적 셀 서식
        area_cell = ws.cell(row=i + 1, column=AREA_COL)
        if isinstance(area_cell.value, (int, float)):
            area_cell.number_format = '#,##0.00'
            area_cell.alignment = right_align
        # 가격 셀 서식
        for offset in range(len(years)):
            cell = ws.cell(row=i + 1, column=PRICE_START_COL + offset)
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'
                cell.alignment = right_align

    widths = [6, 36, 36, 22, 10, 18, 14] + [20] * len(years) + [28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 24
    ws.freeze_panes = "B2"
    wb.save(output_path)
